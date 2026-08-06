"""
Historical backtest of the pattern's forward-return edge.

For every symbol we walk the *entire* available history once and enumerate
every retest -> re-accumulation chain (not just the most recent one, unlike
the live scanner). Each chain resolves to exactly one outcome:

    BOS2_CONFIRMED  - price broke back above P1 with a volume kick
                      (the "FRESH_BOS2" trigger)
    INVALIDATED     - price closed below P0 by more than `max_undercut_pct`
                      before ever breaking out (structure failed)
    TIMEOUT         - re-accumulation dragged on past `max_reaccum_bars`
                      without resolving either way
    STILL_OPEN      - hit the end of available data before resolving
                      (excluded from stats - outcome unknown)

For every chain we also record whether/when it first satisfied the
PRE_BOS2_READY condition (coiled + re-accumulation minimum met), so we can
directly compare two entry strategies:

    1) Enter on BOS2 confirmation (the conservative, confirmed breakout)
    2) Enter on first PRE_BOS2_READY flag (the anticipatory / better-price
       entry the live scanner also raises as its own alert stage)

Both are evaluated with the same realistic mechanics: entry at next bar's
Open, a hard stop at the chain's retest low, and returns measured at fixed
forward horizons.

Every exported row (whether in "All Chains", "BOS2 Trades" or
"PreBOS2 Trades") carries the FULL stage history of that pattern instance -
symbol, P0/base date & price, BOS1 date & price, P1 date & price, retest
date & price, how many bars it re-accumulated, the PRE_BOS2_READY date (if
any), the BOS2 date & price (if confirmed), and the final outcome - so you
can trace exactly which real dates the algorithm used for every trade.

This is research tooling, not investment advice - use it to sanity check
(and re-tune) the scanner's parameters before trusting its live alerts.
"""
from dataclasses import dataclass
from typing import List, Optional

import numpy as np
import pandas as pd

from .pivots import find_reaccum_reversals
from .weekly import find_daily_bos1_candidates

HORIZONS = (1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 12, 15, 20, 25, 30, 40, 60)

# ------------------------------------------------------------------------
# The ONE deliverable: results/backtest_results.xlsx (single file, single
# worksheet). Per explicit user direction (2026-08-08): no more separate
# CSVs/markdown/stop-loss-optimizer files cluttering the repo on every
# backtest run - just this one workbook, built directly from the in-memory
# `all_chains` DataFrame (no intermediate CSV round-trip required).
# ------------------------------------------------------------------------
RESULTS_XLSX_COLUMNS = [
    "Symbol", "Quality Score", "Quality Grade", "Fresh Reversal Entry Date",
    "Fresh Reversal Entry Price", "Re-Accumulation Date", "Retest Date",
    "BOS1 Breakout Date", "PRE_BOS2_READY (Fresh-Entry) Date", "SL Level",
    "Target Price",
]
RESULTS_XLSX_DATE_COLUMNS = [
    "Fresh Reversal Entry Date", "Re-Accumulation Date", "Retest Date",
    "BOS1 Breakout Date", "PRE_BOS2_READY (Fresh-Entry) Date",
]


def build_results_table(all_chains_df: pd.DataFrame, cfg=None) -> pd.DataFrame:
    """Pure transform: raw chain rows (as returned in result["all_chains"],
    or read back from backtest_all_chains.csv) -> the single formatted
    table that goes into backtest_results.xlsx.

    Only keeps chains that reached a confirmed fresh reversal (that's what
    SL Level / Target Price are computed against), de-duped by (symbol,
    reversal date), sorted most-recent-first.
    """
    target_reward_risk = getattr(cfg, "target_reward_risk", 1.0) if cfg is not None else 1.0

    if all_chains_df is None or all_chains_df.empty:
        return pd.DataFrame(columns=RESULTS_XLSX_COLUMNS)

    df = all_chains_df.copy()
    for col in ("bos1_date", "retest_date", "reaccumulation_start_date",
                "reaccum_reversal_date", "pre_bos2_ready_date"):
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors="coerce")

    df = df[df["reaccum_reversal_date"].notna()].copy()
    if df.empty:
        return pd.DataFrame(columns=RESULTS_XLSX_COLUMNS)

    sl_level = df["retest_price"]
    target_price = (df["reaccum_reversal_price"] + target_reward_risk * (df["reaccum_reversal_price"] - sl_level)).round(2)

    out = pd.DataFrame({
        "Symbol": df["symbol"],
        "Quality Score": df["quality_score"],
        "Quality Grade": df["quality_grade"],
        "Fresh Reversal Entry Date": df["reaccum_reversal_date"],
        "Fresh Reversal Entry Price": df["reaccum_reversal_price"],
        "Re-Accumulation Date": df["reaccumulation_start_date"],
        "Retest Date": df["retest_date"],
        "BOS1 Breakout Date": df["bos1_date"],
        "PRE_BOS2_READY (Fresh-Entry) Date": df["pre_bos2_ready_date"],
        "SL Level": sl_level,
        "Target Price": target_price,
    })[RESULTS_XLSX_COLUMNS]

    out = out.drop_duplicates(subset=["Symbol", "Fresh Reversal Entry Date"], keep="first")
    out = out.sort_values("Fresh Reversal Entry Date", ascending=False, na_position="last").reset_index(drop=True)

    for col in RESULTS_XLSX_DATE_COLUMNS:
        out[col] = pd.to_datetime(out[col]).dt.strftime("%Y-%m-%d").replace("NaT", "")

    return out


def write_results_excel(out: pd.DataFrame, out_path: str) -> str:
    """Style + save the single-sheet workbook built by build_results_table()."""
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl import load_workbook

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    quality_fill = {"A": "C6EFCE", "B": "DDEBF7", "C": "FFEB9C", "D": "FFC7CE"}

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        out.to_excel(writer, sheet_name="Backtest Results", index=False)

    wb = load_workbook(out_path)
    ws = wb["Backtest Results"]
    ws.freeze_panes = "A2"

    for col_idx, col in enumerate(out.columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        max_len = max([len(str(col))] + [len(str(v)) for v in out[col].astype(str).values[:1000]])
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 34)

    if "Quality Grade" in out.columns and len(out):
        quality_col_idx = out.columns.get_loc("Quality Grade") + 1
        for row_idx in range(2, len(out) + 2):
            val = ws.cell(row=row_idx, column=quality_col_idx).value
            letter = str(val)[0] if val else None
            fill_hex = quality_fill.get(letter)
            if fill_hex:
                ws.cell(row=row_idx, column=quality_col_idx).fill = PatternFill(
                    start_color=fill_hex, end_color=fill_hex, fill_type="solid")

    ref = f"A1:{get_column_letter(len(out.columns))}{max(len(out) + 1, 1)}"
    tbl = Table(displayName="BacktestResults", ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(tbl)

    wb.save(out_path)
    return out_path



@dataclass
class Chain:
    symbol: str
    p0_date: object
    p0_price: float
    bos1_idx: int
    bos1_price: float
    p1_idx: int
    p1_price: float
    retest_idx: int
    retest_price: float
    outcome: str
    end_idx: Optional[int]
    pre_bos2_ready_idx: Optional[int] = None


def enumerate_chains(df: pd.DataFrame, cfg, symbol: str) -> List[Chain]:
    n = len(df)

    close = df["Close"].values
    high = df["High"].values
    low = df["Low"].values
    vol = df["Volume"].values
    vol_sma20 = df["VOL_SMA20"].values if "VOL_SMA20" in df.columns else np.full(n, np.nan)

    seen_keys = set()
    chains: List[Chain] = []

    bos1_candidates = find_daily_bos1_candidates(df, cfg)

    for bos1_idx, p0_price, p0_date in bos1_candidates:
        run_max_idx, run_max, p1_idx = bos1_idx, high[bos1_idx], None
        k = bos1_idx + 1
        while k < n:
            if high[k] > run_max:
                run_max, run_max_idx = high[k], k
            if close[k] < run_max * 0.97:
                p1_idx = run_max_idx
                break
            k += 1

        if p1_idx is None:
            continue
        p1_price = high[p1_idx]

        retest_idx, invalidated = None, False
        m, running_min, running_min_idx = p1_idx + 1, np.inf, None
        while m < n:
            if low[m] < running_min:
                running_min, running_min_idx = low[m], m
            if close[m] < p0_price * (1 - cfg.max_undercut_pct):
                invalidated = True
                break
            zone_lo, zone_hi = p0_price * (1 - cfg.retest_zone_pct), p0_price * (1 + cfg.retest_zone_pct)
            if running_min_idx is not None and zone_lo <= running_min <= zone_hi:
                if close[m] > running_min * 1.02:
                    retest_idx = running_min_idx
                    break
            m += 1
        if invalidated or retest_idx is None:
            continue

        key = (bos1_idx, retest_idx)
        if key in seen_keys:
            continue
        seen_keys.add(key)
        retest_price = low[retest_idx]

        earliest_bos2 = retest_idx + cfg.min_reaccum_bars
        latest_bos2 = retest_idx + cfg.max_reaccum_bars
        outcome, end_idx, pre_ready_idx = "STILL_OPEN", None, None

        for r in range(retest_idx + 1, min(latest_bos2, n)):
            if close[r] < p0_price * (1 - cfg.max_undercut_pct):
                outcome, end_idx = "INVALIDATED", r
                break
            if r >= earliest_bos2:
                dist = (p1_price - close[r]) / p1_price
                if pre_ready_idx is None and dist <= cfg.pre_bos2_proximity_pct and close[r] < p1_price * (1 + cfg.breakout_buffer):
                    pre_ready_idx = r
                vsma = vol_sma20[r] if not np.isnan(vol_sma20[r]) else np.inf
                if close[r] > p1_price * (1 + cfg.breakout_buffer) and vol[r] >= cfg.vol_mult_bos2 * vsma:
                    outcome, end_idx = "BOS2_CONFIRMED", r
                    break
        else:
            if min(latest_bos2, n) <= n - 1 and latest_bos2 <= n:
                outcome, end_idx = "TIMEOUT", min(latest_bos2, n) - 1

        chains.append(Chain(
            symbol=symbol, p0_date=p0_date, p0_price=p0_price, bos1_idx=bos1_idx, bos1_price=close[bos1_idx],
            p1_idx=p1_idx, p1_price=p1_price, retest_idx=retest_idx, retest_price=retest_price,
            outcome=outcome, end_idx=end_idx, pre_bos2_ready_idx=pre_ready_idx,
        ))

    return chains


def _d(df, idx):
    return df.index[idx] if idx is not None else None


def chain_base_info(chain: Chain, df: pd.DataFrame, cfg=None) -> dict:
    """Every stage date/price for a chain, human-readable, regardless of outcome."""
    from .indicators import confluence_score
    from .scoring import compute_quality_score

    reaccum_end_idx = chain.end_idx if chain.outcome != "STILL_OPEN" else len(df) - 1
    left, right = (cfg.pivot_left, cfg.pivot_right) if cfg is not None else (3, 3)
    reversals = find_reaccum_reversals(df, chain.retest_idx, reaccum_end_idx, left, right)
    last_reversal = reversals[-1] if reversals else (None, None, None)

    # score as-of the most relevant signal bar (the reversal if one formed,
    # else the retest) so historical rows are graded the same way a live
    # scan would have graded them at that point in time
    signal_idx = last_reversal[0] if last_reversal[0] is not None else chain.retest_idx
    conf = confluence_score(df.iloc[: signal_idx + 1], cfg) if cfg is not None else None
    stage_equiv = {
        "BOS2_CONFIRMED": "FRESH_BOS2", "INVALIDATED": "IN_RETEST",
        "TIMEOUT": "BASING", "STILL_OPEN": "BASING",
    }.get(chain.outcome, "BASING")
    if last_reversal[0] is not None:
        stage_equiv = "FRESH_REVERSAL" if chain.outcome != "BOS2_CONFIRMED" else "FRESH_BOS2"
    quality = compute_quality_score(
        stage=stage_equiv,
        confluence_raw=conf["score"] if conf else None, confluence_max=conf["max_score"] if conf else 9,
        num_reversals=len(reversals),
        volatility_contracted=None,
        reaccum_bars=(reaccum_end_idx - chain.retest_idx),
    )

    info = {
        "symbol": chain.symbol,
        "outcome": chain.outcome,
        "quality_score": quality["quality_score"],
        "quality_grade": quality["quality_grade"],
        "p0_date": chain.p0_date,
        "p0_price": round(float(chain.p0_price), 2),
        "bos1_date": _d(df, chain.bos1_idx),
        "bos1_price": round(float(chain.bos1_price), 2),
        "p1_date": _d(df, chain.p1_idx),
        "p1_price": round(float(chain.p1_price), 2),
        "retest_date": _d(df, chain.retest_idx),
        "retest_price": round(float(chain.retest_price), 2),
        "reaccumulation_start_date": _d(df, chain.retest_idx + 1) if chain.retest_idx + 1 < len(df) else None,
        "reaccum_reversal_date": last_reversal[1],
        "reaccum_reversal_price": last_reversal[2],
        "num_reaccum_reversals": len(reversals),
        "pre_bos2_ready_date": _d(df, chain.pre_bos2_ready_idx),
        "reaccum_bars": (chain.end_idx - chain.retest_idx) if chain.outcome != "STILL_OPEN" else None,
        "bos2_date": _d(df, chain.end_idx) if chain.outcome == "BOS2_CONFIRMED" else None,
        "bos2_price": round(float(df["Close"].values[chain.end_idx]), 2) if chain.outcome == "BOS2_CONFIRMED" else None,
        "invalidated_date": _d(df, chain.end_idx) if chain.outcome == "INVALIDATED" else None,
        "timeout_date": _d(df, chain.end_idx) if chain.outcome == "TIMEOUT" else None,
    }
    return info



def _forward_trade(df: pd.DataFrame, entry_idx: int, stop_price: float, horizons=HORIZONS) -> Optional[dict]:

    """
    Simulates one trade: enter at the next bar's Open after `entry_idx`, hard
    stop at `stop_price`, measure returns at each horizon in `horizons`.

    Three possible outcomes per row, captured in `trade_status`:
      PENDING_ENTRY - the signal fired on the most recent available bar, so
                      there's no "next session" data yet to actually enter on
                      (this is what a signal generated *today* looks like -
                      it's still included in the export, just with entry/
                      returns left blank until the next session's data exists)
      STOPPED_OUT   - the stop was hit at some point; every horizon from
                      then on reports that same locked-in exit return, even
                      if we don't yet have that many days of trailing data
                      (once stopped, the trade is closed - no more data needed)
      OPEN/COMPLETE - never stopped; a horizon shows the mark-to-close return
                      if we have that much trailing data yet, else blank
    """
    n = len(df)
    if entry_idx >= n - 1:
        return {
            "entry_date": None, "entry_price": None, "stop_price": round(float(stop_price), 2),
            **{f"ret_{h}d_pct": None for h in horizons},
            "stopped_out": None, "stopped_out_within_days": None,
            "trade_status": "PENDING_ENTRY (signal is on the most recent bar - no next-session open yet)",
        }

    entry_price = df["Open"].values[entry_idx + 1]
    if entry_price <= 0 or np.isnan(entry_price):
        return None

    close = df["Close"].values
    low = df["Low"].values
    result = {"entry_date": df.index[entry_idx + 1], "entry_price": round(float(entry_price), 2),
              "stop_price": round(float(stop_price), 2)}

    stopped_at = None
    for h in horizons:
        if stopped_at is not None:
            # trade already closed at the stop earlier - that return is final
            # regardless of whether `h` days of trailing data exist yet
            result[f"ret_{h}d_pct"] = round(float((stop_price - entry_price) / entry_price * 100), 2)
            continue
        end_i = entry_idx + 1 + h
        if end_i >= n:
            result[f"ret_{h}d_pct"] = None
            continue
        window_low = low[entry_idx + 1: end_i + 1].min()
        if window_low <= stop_price:
            stopped_at = h
            ret = (stop_price - entry_price) / entry_price * 100
        else:
            ret = (close[end_i] - entry_price) / entry_price * 100
        result[f"ret_{h}d_pct"] = round(float(ret), 2)

    result["stopped_out"] = stopped_at is not None
    result["stopped_out_within_days"] = stopped_at
    last_horizon_available = entry_idx + 1 + horizons[-1] < n
    result["trade_status"] = "STOPPED_OUT" if stopped_at is not None else (
        "COMPLETE" if last_horizon_available else "OPEN (still within holding period, more horizons pending)")
    return result


# canonical column order for the trade-level exports
TRADE_COLUMN_ORDER = [
    "symbol", "outcome", "quality_score", "quality_grade", "trade_status",
    "p0_date", "p0_price", "bos1_date", "bos1_price", "p1_date", "p1_price",
    "retest_date", "retest_price", "reaccumulation_start_date", "reaccum_bars",
    "reaccum_reversal_date", "reaccum_reversal_price", "num_reaccum_reversals",
    "pre_bos2_ready_date", "bos2_date", "bos2_price",
    "signal_date", "signal_price", "entry_date", "entry_price", "stop_price",
    "ret_1d_pct", "ret_2d_pct", "ret_3d_pct", "ret_4d_pct", "ret_5d_pct", "ret_6d_pct",
    "ret_7d_pct", "ret_8d_pct", "ret_9d_pct", "ret_10d_pct", "ret_12d_pct", "ret_15d_pct",
    "ret_20d_pct", "ret_25d_pct", "ret_30d_pct", "ret_40d_pct", "ret_60d_pct",

    "stopped_out", "stopped_out_within_days", "eventually_confirmed",
]


def _reorder(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    cols = [c for c in TRADE_COLUMN_ORDER if c in df.columns] + [c for c in df.columns if c not in TRADE_COLUMN_ORDER]
    return df[cols]


def _process_symbol(symbol: str, cfg, data_source, horizons, days: int = None) -> Optional[dict]:
    """Fetch + fully process one symbol's backtest chains and live-signal check.
    Pure function (no shared state) so it's safe to run across threads."""
    from .indicators import add_indicators, confluence_score
    from .scoring import compute_quality_score
    from .trade_plan import compute_trade_plan
    from .pattern import detect_pattern

    try:
        df = data_source.fetch_ohlc(symbol, days=days)
    except Exception as e:
        print(f"  [!] {symbol}: {e}")
        return None
    if df is None or df.empty or len(df) < 80:
        return None
    df = add_indicators(df, cfg)
    chains = enumerate_chains(df, cfg, symbol)

    out = {"chain_rows": [], "trades_bos2": [], "trades_pre": [], "trades_reversal": [],
           "live_signal": None, "outcomes": {}}

    # Same-day cross-check with the live scanner's detector: is this
    # symbol showing a PRE_BOS2_READY / FRESH_BOS2 setup as of the very
    # last bar in the data we just fetched? This is what surfaces
    # "PGIL-right-now"-type setups distinctly from the historical stats.
    try:
        live_match = detect_pattern(df, cfg, symbol)
        if live_match is not None and live_match.stage in ("FRESH_BOS2", "FRESH_REVERSAL", "PRE_BOS2_READY"):
            conf = confluence_score(df, cfg)
            quality = compute_quality_score(
                stage=live_match.stage,
                confluence_raw=conf["score"], confluence_max=conf["max_score"],
                num_reversals=live_match.num_reversals,
                volatility_contracted=live_match.volatility_contracted if not pd.isna(live_match.volatility_contracted) else None,
                reaccum_bars=live_match.reaccum_bars,
            )
            trigger_price = live_match.reversal_price if live_match.stage == "FRESH_REVERSAL" and not pd.isna(live_match.reversal_price) else live_match.last_close
            plan = compute_trade_plan(live_match.stage, trigger_price, live_match.last_date, live_match.retest_price, cfg)
            out["live_signal"] = {
                "symbol": symbol, "stage": live_match.stage,
                "quality_score": quality["quality_score"], "quality_grade": quality["quality_grade"],
                "last_date": live_match.last_date, "last_close": round(float(live_match.last_close), 2),
                "p0_price": round(float(live_match.p0_price), 2) if live_match.p0_price else None,
                "bos1_date": live_match.bos1_date, "bos1_price": round(float(live_match.bos1_price), 2) if live_match.bos1_price else None,
                "p1_resistance": round(float(live_match.p1_price), 2) if live_match.p1_price else None,
                "retest_date": live_match.retest_date, "retest_price": round(float(live_match.retest_price), 2) if live_match.retest_price else None,
                "reaccum_bars": live_match.reaccum_bars,
                "reversal_date": live_match.reversal_date,
                "reversal_price": round(float(live_match.reversal_price), 2) if live_match.reversal_price and not pd.isna(live_match.reversal_price) else None,
                "bos2_date": live_match.bos2_date, "bos2_price": round(float(live_match.bos2_price), 2) if live_match.bos2_price else None,
                "entry_date": plan["entry_date"] if plan else None,
                "entry_price_ref": plan["entry_price_ref"] if plan else None,
                "stop_loss": plan["stop_loss"] if plan else None,
                "risk_pct": plan["risk_pct"] if plan else None,
                "target_price": plan["target_price"] if plan else None,
                "target_pct": plan["target_pct"] if plan else None,
                "reward_risk": plan["reward_risk"] if plan else None,
                "exit_by_min_date": plan["exit_by_min_date"] if plan else None,
                "exit_by_max_date": plan["exit_by_max_date"] if plan else None,
                "confluence_score": f"{conf['score']}/{conf['max_score']}",
                "notes": live_match.notes,
            }
    except Exception as e:
        print(f"  [!] {symbol}: live-signal check failed: {e}")

    for c in chains:
        out["outcomes"][c.outcome] = out["outcomes"].get(c.outcome, 0) + 1
        base_info = chain_base_info(c, df, cfg)
        out["chain_rows"].append(base_info)

        if c.outcome == "BOS2_CONFIRMED":
            trade = _forward_trade(df, c.end_idx, c.retest_price, horizons)
            if trade:
                row = dict(base_info)
                row.update({"signal_date": base_info["bos2_date"], "signal_price": base_info["bos2_price"]})
                row.update(trade)
                out["trades_bos2"].append(row)

        if c.pre_bos2_ready_idx is not None:
            trade = _forward_trade(df, c.pre_bos2_ready_idx, c.retest_price, horizons)
            if trade:
                row = dict(base_info)
                row.update({
                    "signal_date": base_info["pre_bos2_ready_date"],
                    "signal_price": round(float(df["Close"].values[c.pre_bos2_ready_idx]), 2),
                    "eventually_confirmed": c.outcome == "BOS2_CONFIRMED",
                })
                row.update(trade)
                out["trades_pre"].append(row)

        # Tactical "higher-low reversal" entry: first green candle closing
        # back above the most recent swing-low candle's high, inside the
        # re-accumulation window - an earlier/tighter entry than waiting
        # for the full P1 breakout.
        if base_info.get("reaccum_reversal_date") is not None:
            rev_idx = df.index.get_loc(base_info["reaccum_reversal_date"])
            trade = _forward_trade(df, rev_idx, c.retest_price, horizons)
            if trade:
                row = dict(base_info)
                row.update({
                    "signal_date": base_info["reaccum_reversal_date"],
                    "signal_price": base_info["reaccum_reversal_price"],
                    "eventually_confirmed": c.outcome == "BOS2_CONFIRMED",
                })
                row.update(trade)
                out["trades_reversal"].append(row)

    return out


def run_backtest(symbols: List[str], cfg, data_source, horizons=HORIZONS, days: int = None,
                  max_workers: int = 8) -> dict:
    """Run the full historical backtest across `symbols`.

    `days`: how many calendar days of history to fetch per symbol. Defaults
    to `cfg.backtest_history_years` (5 years) when not explicitly given -
    see the CLI's `--years` flag.
    `max_workers`: symbols are fetched/processed in parallel threads (I/O
    bound - mostly waiting on Dhan/yfinance) since the default universe is
    now "every symbol >= min_market_cap_cr", not a small curated sample.
    """
    import concurrent.futures as cf

    if days is None:
        days = int(cfg.backtest_history_years * 365.25)

    all_chain_rows = []
    all_trades_bos2 = []
    all_trades_pre = []
    all_trades_reversal = []
    live_signals = []
    outcome_counts = {"BOS2_CONFIRMED": 0, "INVALIDATED": 0, "TIMEOUT": 0, "STILL_OPEN": 0}

    total = len(symbols)
    done = 0
    with cf.ThreadPoolExecutor(max_workers=max_workers) as ex:
        futures = {ex.submit(_process_symbol, sym, cfg, data_source, horizons, days): sym for sym in symbols}
        for fut in cf.as_completed(futures):
            sym = futures[fut]
            done += 1
            try:
                res = fut.result()
            except Exception as e:
                print(f"  [!] {sym}: {e}")
                res = None
            if done % 50 == 0 or done == total:
                print(f"[backtest] processed {done}/{total}")
            if res is None:
                continue
            all_chain_rows.extend(res["chain_rows"])
            all_trades_bos2.extend(res["trades_bos2"])
            all_trades_pre.extend(res["trades_pre"])
            all_trades_reversal.extend(res["trades_reversal"])
            if res["live_signal"] is not None:
                live_signals.append(res["live_signal"])
            for k, v in res["outcomes"].items():
                outcome_counts[k] = outcome_counts.get(k, 0) + v


    all_chains_df = pd.DataFrame(all_chain_rows)
    bos2_df = _reorder(pd.DataFrame(all_trades_bos2))
    pre_df = _reorder(pd.DataFrame(all_trades_pre))
    reversal_df = _reorder(pd.DataFrame(all_trades_reversal))

    live_signals_df = pd.DataFrame(live_signals)
    if not live_signals_df.empty:
        stage_prio = {"FRESH_BOS2": 0, "FRESH_REVERSAL": 1, "PRE_BOS2_READY": 2}
        live_signals_df["_p"] = live_signals_df["stage"].map(stage_prio).fillna(9)
        live_signals_df = live_signals_df.sort_values(["_p", "quality_score"], ascending=[True, False]).drop(columns="_p").reset_index(drop=True)

    summary = _summarize(bos2_df, pre_df, reversal_df, outcome_counts, horizons)
    return {
        "all_chains": all_chains_df,
        "bos2_trades": bos2_df,
        "pre_bos2_trades": pre_df,
        "reversal_trades": reversal_df,
        "live_signals": live_signals_df,
        "outcome_counts": outcome_counts,
        "summary": summary,
    }



def _summarize(bos2_df, pre_df, reversal_df, outcome_counts, horizons) -> pd.DataFrame:
    rows = []

    total_resolved = sum(v for k, v in outcome_counts.items() if k != "STILL_OPEN")
    confirm_rate = outcome_counts.get("BOS2_CONFIRMED", 0) / total_resolved * 100 if total_resolved else float("nan")

    for label, tdf in (("BOS2_CONFIRMED entry", bos2_df), ("PRE_BOS2_READY entry", pre_df),
                       ("Reaccum-Reversal entry", reversal_df)):

        if tdf is None or tdf.empty:
            continue
        for h in horizons:
            col = f"ret_{h}d_pct"
            if col not in tdf.columns:
                continue
            vals = tdf[col].dropna()
            if vals.empty:
                continue
            win_rate = (vals > 0).mean() * 100
            rows.append({
                "entry_type": label, "horizon_days": h, "n_trades": len(vals),
                "win_rate_pct": round(win_rate, 1), "avg_return_pct": round(vals.mean(), 2),
                "median_return_pct": round(vals.median(), 2), "best_pct": round(vals.max(), 2),
                "worst_pct": round(vals.min(), 2),
            })
    summary = pd.DataFrame(rows)
    summary.attrs["pattern_completion_rate_pct"] = round(confirm_rate, 1)
    summary.attrs["outcome_counts"] = outcome_counts
    return summary


def render_markdown_report(result: dict, cfg) -> str:
    summary = result["summary"]
    counts = result["outcome_counts"]
    total_resolved = sum(v for k, v in counts.items() if k != "STILL_OPEN")
    completion_rate = counts.get("BOS2_CONFIRMED", 0) / total_resolved * 100 if total_resolved else float("nan")

    lines = ["# SMC Structure Scanner - Backtest Report", ""]
    lines.append(f"Generated: {pd.Timestamp.now().isoformat(timespec='seconds')}")
    lines.append("")
    lines.append("## Pattern completion rate")
    lines.append("")
    lines.append(f"- Total resolved retest/re-accumulation chains: **{total_resolved}**")
    for k, v in counts.items():
        lines.append(f"  - {k}: {v}")
    lines.append(f"- **{completion_rate:.1f}%** of chains that reached re-accumulation went on to a confirmed BOS2 breakout.")
    lines.append("")
    lines.append("## Forward returns by entry type & horizon")
    lines.append("")
    if not summary.empty:
        lines.append(summary.to_markdown(index=False))
    else:
        lines.append("_No trades generated - widen the symbol universe or date range._")
    lines.append("")
    lines.append("## Notes / caveats")
    lines.append("- Entries assume next-bar Open after the signal; stop = the chain's retest low.")
    lines.append("- No slippage, brokerage, or position sizing modeled.")
    lines.append("- `PRE_BOS2_READY` entries include chains that *never* confirmed BOS2 - that's the "
                  "real cost of entering early, weigh it against the better average price.")
    lines.append("- This console report is not written to disk - the one persisted deliverable is "
                  "`results/backtest_results.xlsx` (single sheet: Symbol, Quality Score/Grade, Fresh "
                  "Reversal Entry Date/Price, Re-Accumulation Date, Retest Date, BOS1 Breakout Date, "
                  "PRE_BOS2_READY (Fresh-Entry) Date, SL Level, Target Price).")
    return "\n".join(lines)
