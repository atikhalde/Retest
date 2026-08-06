"""
Exports the backtest CSV/markdown outputs into one polished, multi-sheet
Excel workbook: results/backtest_results.xlsx

Sheets:
    Summary          headline stats + forward-return table by entry type/horizon
    Outcome Counts    chain resolution breakdown (BOS2_CONFIRMED/INVALIDATED/...)
    By Symbol         per-symbol chain counts & average confirmed-BOS2 returns
    BOS2 Trades       full trade log for the confirmed-breakout entry style
    PreBOS2 Trades    full trade log for the anticipatory pre-breakout entry style
    Notes             caveats / methodology (verbatim from backtest_report.md)
"""
import os
import sys

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.worksheet.table import Table, TableStyleInfo

RESULTS_DIR = os.path.join(os.path.dirname(__file__), "..", "results")

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _style_sheet(ws, df, freeze="A2", as_table=True, table_name=None):
    ws.freeze_panes = freeze
    for col_idx, col in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        max_len = max([len(str(col))] + [len(str(v)) for v in df[col].astype(str).values[:500]])
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 40)

    if as_table and len(df) > 0:
        n_rows, n_cols = df.shape
        last_col = get_column_letter(n_cols)
        ref = f"A1:{last_col}{n_rows + 1}"
        tbl = Table(displayName=table_name or ws.title.replace(" ", "_"), ref=ref)
        tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        ws.add_table(tbl)


def _add_return_color_scale(ws, df, cols):
    for col_name in cols:
        if col_name not in df.columns:
            continue
        col_idx = df.columns.get_loc(col_name) + 1
        col_letter = get_column_letter(col_idx)
        rng = f"{col_letter}2:{col_letter}{len(df) + 1}"
        rule = ColorScaleRule(
            start_type="min", start_color="F8696B",
            mid_type="num", mid_value=0, mid_color="FFEB84",
            end_type="max", end_color="63BE7B",
        )
        ws.conditional_formatting.add(rng, rule)


def build_workbook(results_dir=RESULTS_DIR, out_path=None):
    out_path = out_path or os.path.join(results_dir, "backtest_results.xlsx")

    summary = pd.read_csv(os.path.join(results_dir, "backtest_summary.csv"))
    bos2 = pd.read_csv(os.path.join(results_dir, "backtest_bos2_trades.csv"))
    pre = pd.read_csv(os.path.join(results_dir, "backtest_pre_bos2_trades.csv"))
    all_chains_path = os.path.join(results_dir, "backtest_all_chains.csv")
    all_chains = pd.read_csv(all_chains_path) if os.path.exists(all_chains_path) else pd.DataFrame()

    report_path = os.path.join(results_dir, "backtest_report.md")
    report_text = open(report_path).read() if os.path.exists(report_path) else ""

    # ---- outcome counts (parsed back out of the markdown report header) ----
    import re
    counts = {}
    for m in re.finditer(r"^\s*-\s+(\w+):\s+(\d+)\s*$", report_text, re.MULTILINE):
        counts[m.group(1)] = int(m.group(2))
    outcome_df = pd.DataFrame(
        [{"outcome": k, "count": v} for k, v in counts.items()]
    ) if counts else pd.DataFrame(columns=["outcome", "count"])

    completion_match = re.search(r"\*\*([\d.]+)%\*\* of chains", report_text)
    completion_rate = float(completion_match.group(1)) if completion_match else None

    # ---- per-symbol breakdown from the BOS2-confirmed trade log ----
    by_symbol = pd.DataFrame()
    if not bos2.empty:
        agg = {"symbol": "count"}
        ret_cols = [c for c in bos2.columns if c.startswith("ret_")]
        g = bos2.groupby("symbol")
        rows = []
        for sym, sub in g:
            row = {"symbol": sym, "n_bos2_signals": len(sub)}
            for c in ret_cols:
                vals = sub[c].dropna()
                row[f"{c}_avg"] = round(vals.mean(), 2) if len(vals) else None
                row[f"{c}_winrate"] = round((vals > 0).mean() * 100, 1) if len(vals) else None
            rows.append(row)
        by_symbol = pd.DataFrame(rows).sort_values("n_bos2_signals", ascending=False)

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        # ---- Summary sheet ----
        meta_rows = [
            {"metric": "Report generated", "value": pd.Timestamp.now().isoformat(timespec="seconds")},
            {"metric": "Total resolved chains", "value": int(outcome_df["count"].sum()) if not outcome_df.empty else None},
            {"metric": "Pattern completion rate (%)", "value": completion_rate},
            {"metric": "All pattern instances (incl. invalidated/timeout)", "value": len(all_chains)},
            {"metric": "BOS2-confirmed trades", "value": len(bos2)},
            {"metric": "PRE_BOS2_READY trades", "value": len(pre)},
        ]
        meta_df = pd.DataFrame(meta_rows)
        meta_df.to_excel(writer, sheet_name="Summary", index=False, startrow=0)
        summary.to_excel(writer, sheet_name="Summary", index=False, startrow=len(meta_df) + 3)

        outcome_df.to_excel(writer, sheet_name="Outcome Counts", index=False)
        if not by_symbol.empty:
            by_symbol.to_excel(writer, sheet_name="By Symbol", index=False)
        if not all_chains.empty:
            all_chains.to_excel(writer, sheet_name="All Chains", index=False)
        bos2.to_excel(writer, sheet_name="BOS2 Trades", index=False)
        pre.to_excel(writer, sheet_name="PreBOS2 Trades", index=False)


        notes_df = pd.DataFrame({"Backtest report (methodology & caveats)": report_text.split("\n")})
        notes_df.to_excel(writer, sheet_name="Notes", index=False)

    # ---- post-process styling with openpyxl ----
    from openpyxl import load_workbook
    wb = load_workbook(out_path)

    ws = wb["Summary"]
    for cell in ws[1]:
        cell.font = Font(bold=True)
    header_row = len(meta_df) + 4
    for cell in ws[header_row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for col_idx in range(1, summary.shape[1] + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18
    ret_cols_idx = [i for i, c in enumerate(summary.columns) if "return" in c or "rate" in c]
    for i in ret_cols_idx:
        col_letter = get_column_letter(i + 1)
        rng = f"{col_letter}{header_row+1}:{col_letter}{header_row+len(summary)}"
        rule = ColorScaleRule(start_type="min", start_color="F8696B",
                               mid_type="num", mid_value=0, mid_color="FFEB84",
                               end_type="max", end_color="63BE7B")
        ws.conditional_formatting.add(rng, rule)

    if "Outcome Counts" in wb.sheetnames and not outcome_df.empty:
        _style_sheet(wb["Outcome Counts"], outcome_df, table_name="OutcomeCounts")

    if "By Symbol" in wb.sheetnames and not by_symbol.empty:
        ws2 = wb["By Symbol"]
        _style_sheet(ws2, by_symbol, table_name="BySymbol")
        _add_return_color_scale(ws2, by_symbol, [c for c in by_symbol.columns if c.endswith("_avg")])

    if "All Chains" in wb.sheetnames and not all_chains.empty:
        _style_sheet(wb["All Chains"], all_chains, table_name="AllChains")

    if not bos2.empty:
        ws3 = wb["BOS2 Trades"]
        _style_sheet(ws3, bos2, table_name="BOS2Trades")
        _add_return_color_scale(ws3, bos2, [c for c in bos2.columns if c.startswith("ret_")])

    if not pre.empty:
        ws4 = wb["PreBOS2 Trades"]
        _style_sheet(ws4, pre, table_name="PreBOS2Trades")
        _add_return_color_scale(ws4, pre, [c for c in pre.columns if c.startswith("ret_")])

    ws5 = wb["Notes"]
    ws5.column_dimensions["A"].width = 110
    for row in ws5.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(out_path)
    print(f"Wrote {out_path}")
    return out_path


if __name__ == "__main__":
    build_workbook()
