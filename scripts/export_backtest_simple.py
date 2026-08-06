"""
Builds ONE simple, plain-English worksheet from the backtest CSVs:
results/backtest_simple.xlsx

Columns (left to right, in the order things actually happen):
    Stock, Status, Base Date, Base Price, Breakout Date, Breakout Price,
    Peak Date, Peak Price, Retest Date, Retest Price,
    Re-Accumulation Start, Re-Accumulation Days,
    Pre-Breakout Alert Date, Confirmed Breakout Date, Confirmed Breakout Price,
    Entry Date, Entry Price, Stop Loss,
    Return 5D %, Return 10D %, Return 20D %, Return 40D %, Return 60D %,
    Trade Status

One row per pattern instance found (whether or not it ever became a trade),
sorted so the most recently active setups (e.g. today's PGIL-style breakout)
are at the top.

For the full, multi-sheet, everything-broken-out version, run
`export_backtest_excel.py` instead (produces backtest_results.xlsx).
"""
import os

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.worksheet.table import Table, TableStyleInfo

RESULTS_DIR = os.path.join(os.path.dirname(__file__), "..", "results")

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)

STATUS_MAP = {
    "BOS2_CONFIRMED": "Breakout Confirmed",
    "INVALIDATED": "Failed (support broke)",
    "TIMEOUT": "Timed Out (never broke out)",
    "STILL_OPEN": "Still Forming",
}
STATUS_FILL = {
    "Breakout Confirmed": "C6EFCE",       # green
    "Failed (support broke)": "FFC7CE",   # red
    "Timed Out (never broke out)": "FFEB9C",  # yellow
    "Still Forming": "DDEBF7",            # light blue
}
TRADE_STATUS_MAP = {
    "STOPPED_OUT": "Stopped Out",
    "COMPLETE": "Trade Complete",
}

COLUMN_RENAME = {
    "symbol": "Stock",
    "p0_date": "Base Date",
    "p0_price": "Base Price",
    "bos1_date": "Breakout Date",
    "bos1_price": "Breakout Price",
    "p1_date": "Peak Date",
    "p1_price": "Peak Price",
    "retest_date": "Retest Date",
    "retest_price": "Retest Price",
    "reaccumulation_start_date": "Re-Accumulation Start",
    "reaccum_bars": "Re-Accumulation Days",
    "pre_bos2_ready_date": "Pre-Breakout Alert Date",
    "bos2_date": "Confirmed Breakout Date",
    "bos2_price": "Confirmed Breakout Price",
    "entry_date": "Entry Date",
    "entry_price": "Entry Price",
    "stop_price": "Stop Loss",
    "ret_5d_pct": "Return 5D %",
    "ret_10d_pct": "Return 10D %",
    "ret_20d_pct": "Return 20D %",
    "ret_40d_pct": "Return 40D %",
    "ret_60d_pct": "Return 60D %",
}

FINAL_COLUMN_ORDER = [
    "Stock", "Status",
    "Base Date", "Base Price",
    "Breakout Date", "Breakout Price",
    "Peak Date", "Peak Price",
    "Retest Date", "Retest Price",
    "Re-Accumulation Start", "Re-Accumulation Days",
    "Pre-Breakout Alert Date",
    "Confirmed Breakout Date", "Confirmed Breakout Price",
    "Entry Date", "Entry Price", "Stop Loss",
    "Return 5D %", "Return 10D %", "Return 20D %", "Return 40D %", "Return 60D %",
    "Trade Status",
]


def _clean_trade_status(v):
    if pd.isna(v):
        return "-"
    v = str(v)
    if v.startswith("PENDING_ENTRY"):
        return "Pending (triggered on latest bar - no return yet)"
    if v.startswith("OPEN"):
        return "Trade Open (still running)"
    return TRADE_STATUS_MAP.get(v, v)


def build_simple_workbook(results_dir=RESULTS_DIR, out_path=None) -> str:
    out_path = out_path or os.path.join(results_dir, "backtest_simple.xlsx")

    all_chains = pd.read_csv(os.path.join(results_dir, "backtest_all_chains.csv"),
                              parse_dates=["p0_date", "bos1_date", "p1_date", "retest_date",
                                           "reaccumulation_start_date", "pre_bos2_ready_date",
                                           "bos2_date"])
    bos2_path = os.path.join(results_dir, "backtest_bos2_trades.csv")
    bos2 = pd.read_csv(bos2_path, parse_dates=["p0_date", "bos1_date", "retest_date", "entry_date"]) if os.path.exists(bos2_path) else pd.DataFrame()

    key_cols = ["symbol", "p0_date", "bos1_date", "retest_date"]
    trade_cols = ["entry_date", "entry_price", "stop_price",
                  "ret_5d_pct", "ret_10d_pct", "ret_20d_pct", "ret_40d_pct", "ret_60d_pct",
                  "trade_status"]

    if not bos2.empty:
        merged = all_chains.merge(bos2[key_cols + trade_cols], on=key_cols, how="left")
    else:
        merged = all_chains.copy()
        for c in trade_cols:
            merged[c] = None

    merged["Status"] = merged["outcome"].map(STATUS_MAP).fillna(merged["outcome"])
    merged = merged.rename(columns=COLUMN_RENAME)
    merged["Trade Status"] = merged["trade_status"].apply(_clean_trade_status) if "trade_status" in merged.columns else "-"

    for col in FINAL_COLUMN_ORDER:
        if col not in merged.columns:
            merged[col] = None
    out = merged[FINAL_COLUMN_ORDER].copy()

    # most relevant/most recent date first: confirmed breakout > pre-breakout alert > retest
    out["_sort_date"] = out["Confirmed Breakout Date"].fillna(out["Pre-Breakout Alert Date"]).fillna(out["Retest Date"])
    out = out.sort_values("_sort_date", ascending=False).drop(columns="_sort_date").reset_index(drop=True)

    for col in ["Base Date", "Breakout Date", "Peak Date", "Retest Date", "Re-Accumulation Start",
                "Pre-Breakout Alert Date", "Confirmed Breakout Date", "Entry Date"]:
        out[col] = pd.to_datetime(out[col]).dt.strftime("%Y-%m-%d").replace("NaT", "")

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        out.to_excel(writer, sheet_name="Backtest Results", index=False)

    from openpyxl import load_workbook
    wb = load_workbook(out_path)
    ws = wb["Backtest Results"]
    ws.freeze_panes = "C2"

    for col_idx, col in enumerate(out.columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        max_len = max([len(str(col))] + [len(str(v)) for v in out[col].astype(str).values[:1000]])
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 34)

    status_col_idx = out.columns.get_loc("Status") + 1
    for row_idx in range(2, len(out) + 2):
        status_val = ws.cell(row=row_idx, column=status_col_idx).value
        fill_hex = STATUS_FILL.get(status_val)
        if fill_hex:
            fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type="solid")
            for col_idx in range(1, len(out.columns) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = fill

    ret_cols = ["Return 5D %", "Return 10D %", "Return 20D %", "Return 40D %", "Return 60D %"]
    for col_name in ret_cols:
        col_idx = out.columns.get_loc(col_name) + 1
        col_letter = get_column_letter(col_idx)
        rng = f"{col_letter}2:{col_letter}{len(out) + 1}"
        rule = ColorScaleRule(start_type="min", start_color="F8696B",
                               mid_type="num", mid_value=0, mid_color="FFFFFF",
                               end_type="max", end_color="63BE7B")
        ws.conditional_formatting.add(rng, rule)

    ref = f"A1:{get_column_letter(len(out.columns))}{len(out) + 1}"
    tbl = Table(displayName="BacktestResults", ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=False)
    ws.add_table(tbl)

    wb.save(out_path)
    print(f"Wrote {out_path}  ({len(out)} rows)")
    return out_path


if __name__ == "__main__":
    build_simple_workbook()
